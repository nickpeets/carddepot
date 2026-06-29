#!/usr/bin/env python3
"""Re-derive card URLs/teams from data/master_checklist.xlsx and rewrite data/cards-*.json.

Per record decision order:
  (a) base-set card (set==brand) with a strict player match in the brand's
      "Year Brand" tab -> use that row's own URL + Team.
  (b) else, if the existing URL slug ends with the card's OWN player -> keep URL (and team).
  (c) else -> clear URL AND team.
  (d) else (no URL to begin with) -> leave empty.

Hard guard: per-year record counts must be identical before/after (155844 total).
Output preserves compact single-line JSON: separators=(",", ":"), no trailing newline.
"""
import json, glob, io, os, re, sys
from urllib.parse import urlsplit, unquote

DATA = "data"
XLSX = os.path.join(DATA, "master_checklist.xlsx")

# suffix variation tags to strip (case-insensitive token match)
SUFFIX_TAGS = {"FB", "FBC", "HL", "RC", "VAR"}

def norm_name(s):
    """Normalize a player name for comparison."""
    s = (s or "")
    s = s.replace("&", " and ")
    s = re.sub(r"[^A-Za-z0-9 ]+", " ", s)   # drop punctuation/accents-ish
    s = re.sub(r"\s+", " ", s).strip().lower()
    return s

def strip_player(name):
    """Strip suffix variation tags (FB/FBC/HL/RC/VAR) and trailing ALL-CAPS tokens."""
    s = (name or "").strip()
    toks = s.split()
    # repeatedly drop trailing tokens that are known tags OR all-caps (len>=1)
    while toks:
        last = toks[-1]
        bare = re.sub(r"[^A-Za-z]", "", last)
        if not bare:
            toks.pop(); continue
        if last.upper() in SUFFIX_TAGS:
            toks.pop(); continue
        # trailing all-caps token (e.g. DK, IA, AS, UER) -> variation/tag
        if bare.isupper():
            toks.pop(); continue
        break
    return " ".join(toks)

def slug_player_tail(url):
    """Return the normalized trailing player portion of a tcdb-style URL slug."""
    if not url:
        return ""
    try:
        path = urlsplit(url).path
    except Exception:
        path = url
    seg = unquote(path.rstrip("/").split("/")[-1])
    # seg like '1985-Donruss-8-Carney-Lansford'  -> drop 'YYYY-Brand-Num-' prefix
    # find first numeric token; player is everything after it
    parts = seg.split("-")
    idx = None
    for i, p in enumerate(parts):
        if re.fullmatch(r"\d+[A-Za-z]?", p):  # card number token (e.g. 8, 12a)
            idx = i
            break
    tail = parts[idx+1:] if idx is not None else parts
    return norm_name(" ".join(tail))

def load_xlsx():
    import openpyxl
    wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)
    # map: (year, brand) -> { number_str -> {player, team, url} }  (strict-player keyed)
    # We index by normalized stripped player within each (sheet, number) for strict match.
    sheet_idx = {}  # (year,brand) -> { num -> list of rows }
    for name in wb.sheetnames:
        m = re.match(r"^(\d{4})\s+(.+?)\s*$", name)
        if not m:
            continue
        year, brand = m.group(1), m.group(2).strip()
        ws = wb[name]
        rows = ws.iter_rows(min_row=2, values_only=True)
        bucket = {}
        for r in rows:
            if not r:
                continue
            cardno = r[0] if len(r) > 0 else None
            player = r[1] if len(r) > 1 else None
            team   = r[2] if len(r) > 2 else None
            url    = r[3] if len(r) > 3 else None
            if cardno is None:
                continue
            # normalize card number: 8.0 -> '8'
            if isinstance(cardno, float) and cardno.is_integer():
                num = str(int(cardno))
            else:
                num = str(cardno).strip()
            bucket.setdefault(num, []).append({
                "player": player, "team": team, "url": url,
            })
        sheet_idx[(year, brand)] = bucket
    wb.close()
    return sheet_idx

def main():
    dry = "--write" not in sys.argv
    sheet_idx = load_xlsx()
    files = sorted(glob.glob(os.path.join(DATA, "cards-*.json")))
    baseline = json.loads(io.open("/tmp/baseline_counts.json").read()) if os.path.exists("/tmp/baseline_counts.json") else {}
    stats = {"fixed":0, "kept":0, "cleared":0, "empty":0, "total":0}
    out_payload = {}
    count_violation = None
    for f in files:
        year = re.search(r"cards-(\d+)\.json", f).group(1)
        data = json.loads(io.open(f, encoding="utf-8").read())
        n_before = len(data)
        for rec in data:
            stats["total"] += 1
            brand = rec.get("brand") or ""
            cset  = rec.get("set") or ""
            num   = str(rec.get("number") or "").strip()
            player= rec.get("player") or ""
            cur_url = rec.get("url") or ""
            pn = norm_name(strip_player(player))
            decided = False
            # (a) base-set + strict player match
            if cset == brand:
                bucket = sheet_idx.get((year, brand), {})
                cand = bucket.get(num, [])
                hit = None
                for row in cand:
                    if norm_name(strip_player(row.get("player"))) == pn and pn:
                        hit = row; break
                if hit is not None:
                    new_url = hit.get("url") or ""
                    new_team = hit.get("team") or ""
                    rec["url"] = new_url
                    if new_team:
                        rec["team"] = new_team
                    stats["fixed"] += 1
                    decided = True
            if not decided:
                if cur_url:
                    # (b) keep if existing slug ends with this card's OWN player
                    if pn and slug_player_tail(cur_url).endswith(pn):
                        stats["kept"] += 1
                    else:
                        # (c) clear URL and team
                        rec["url"] = ""
                        rec["team"] = ""
                        stats["cleared"] += 1
                else:
                    # (d) leave empty
                    stats["empty"] += 1
        n_after = len(data)
        if n_before != n_after:
            count_violation = (year, n_before, n_after)
            break
        if baseline and str(year) in baseline and baseline[str(year)] != n_after:
            count_violation = (year, baseline[str(year)], n_after)
            break
        out_payload[f] = data
    if count_violation:
        print("COUNT GUARD VIOLATION (year, before, after):", count_violation)
        print("ABORTING - no files written.")
        sys.exit(2)
    have_url = stats["fixed"] + stats["kept"]
    cov = 100.0 * have_url / stats["total"] if stats["total"] else 0
    print("=== SCOPE ===")
    print("total records :", stats["total"])
    print("fixed (a)     :", stats["fixed"])
    print("kept  (b)     :", stats["kept"])
    print("cleared (c)   :", stats["cleared"])
    print("empty (d)     :", stats["empty"])
    print("coverage (have URL) : {:.2f}%".format(cov))
    if dry:
        print("DRY RUN - no files written. Pass --write to persist.")
        return
    for f, data in out_payload.items():
        s = json.dumps(data, ensure_ascii=False, separators=(",", ":"))
        with io.open(f, "w", encoding="utf-8", newline="") as fh:
            fh.write(s)  # no trailing newline
    print("WROTE", len(out_payload), "files (compact, no trailing newline).")

if __name__ == "__main__":
    main()
